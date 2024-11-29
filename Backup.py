import datetime
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import mysql.connector


current_datetime = datetime.datetime.now()

# Extract the date, time, and day
current_date = current_datetime.date()
current_time = current_datetime.time()
current_day = current_datetime.strftime("%A")
whichclass=""
# Print the results
print("Current Date:", current_date)
print("Current Time:", current_time)
print("Current Day:", current_day)

# Convert the time and date to strings
current_time_str = current_time.strftime("%H:%M:%S")
current_date_str = current_date.strftime("%Y-%m-%d")

# Extract the hour and minute from the current time
sumtime = current_time_str[0:2]
subtime = current_time_str[3:5]
# Establish a connection to the MySQL database

# Connect to the MySQL database
connection = mysql.connector.connect(
    host='localhost',
    user='root',
    password='sathvik#123',
    database='genralattendance'
)



firstdigithr = int(sumtime[0])
lastdigithr = int(sumtime[1])

integerhr = (firstdigithr * 10) + lastdigithr

firstdigitmin = int(subtime[0])
lastdigitmin = int(subtime[1])

integermin = (firstdigitmin * 10) + lastdigitmin

column=""

def class1(column):
    column="S"
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    # Load the existing workbook
    print("Class 1 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 1.xlsx")
    

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 1.xlsx")
    print("Attendance registered For class 1")
    return

def class2(column):
    column="O"
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    # Load the existing workbook
    print("Class 2 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 2.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 2.xlsx")
    print("Attendance registered For class 2")
    return

def class3(column):
    column="W"
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    # Load the existing workbook
    print("Class 3 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 3.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 3.xlsx")
    print("Attendance registered For class 3")
    return

def class4(column):
    # Load the existing workbook
    column="O"
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + 1)
    print("Class 4 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 4.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 4.xlsx")
    print("Attendance registered For class 4")
    return

def attendreportclass1():
    try:
        dfclass1=pd.read_excel("G:/python project/Class 1.xlsx")
        cursor = connection.cursor()

        # Convert the DataFrame to a list of tuples
        data = dfclass1.values.tolist()

        # Define the SQL query to insert the data into the table
        table_name = 'Class 1'
        query = f"INSERT INTO {table_name} (column1, column2, ...) VALUES (%s, %s, ...)"
        print("Data added to MySQL")

        # Execute the SQL query for each row in the DataFrame
        cursor.executemany(query, data)
        connection.commit()
        print(dfclass1.to_string())
        return
    except:
        dfclass1=pd.read_excel("G:/python project/Class 1.xlsx")
        print(dfclass1.to_string())
        return
def attendreportclass2():
    try:
        dfclass2=pd.read_excel("G:/python project/Class 2.xlsx")
        cursor = connection.cursor()

        # Convert the DataFrame to a list of tuples
        data = dfclass2.values.tolist()

        # Define the SQL query to insert the data into the table
        table_name = 'Class 2'
        query = f"INSERT INTO {table_name} (column1, column2, ...) VALUES (%s, %s, ...)"
        print("Data added to MySQL")

        # Execute the SQL query for each row in the DataFrame
        cursor.executemany(query, data)
        connection.commit()
        print(dfclass2.to_string())
        return
    except:
        dfclass2=pd.read_excel("G:/python project/Class 2.xlsx")
        print(dfclass2.to_string())
        return

def attendreportclass3():
    try:
        dfclass3=pd.read_excel("G:/python project/Class 3.xlsx")
        cursor = connection.cursor()

        # Convert the DataFrame to a list of tuples
        data = dfclass3.values.tolist()

        # Define the SQL query to insert the data into the table
        table_name = 'Class 3'
        query = f"INSERT INTO {table_name} (column1, column2, ...) VALUES (%s, %s, ...)"
        print("Data added to MySQL")

        # Execute the SQL query for each row in the DataFrame
        cursor.executemany(query, data)
        connection.commit()
        print(dfclass3.to_string())
        return
    except:
        dfclass3=pd.read_excel("G:/python project/Class 3.xlsx")
        print(dfclass3.to_string())
        return

def attendreportclass4():
    try:
        dfclass4=pd.read_excel("G:/python project/Class 4.xlsx")
        cursor = connection.cursor()

        # Convert the DataFrame to a list of tuples
        data = dfclass4.values.tolist()

        # Define the SQL query to insert the data into the table
        table_name = 'Class 4'
        query = f"INSERT INTO {table_name} (column1, column2, ...) VALUES (%s, %s, ...)"
        print("Data added to MySQL")

        # Execute the SQL query for each row in the DataFrame
        cursor.executemany(query, data)
        connection.commit()
        print(dfclass4.to_string())
        return
    except:
        dfclass4=pd.read_excel("G:/python project/Class 4.xlsx")
        print(dfclass4.to_string())
        return

def dateclass1():
    try:
        dfclass1=pd.read_excel("G:/python project/Class 1.xlsx")
        indate=input("\nDate (DD-MM-YYYY) :")
        conroll=dfclass1['Roll no']
        conname=dfclass1['Name']
        condate=dfclass1[indate]
        absent=0
        present=0
        for i in range(0,30,1):
            if(condate[i]=='A'):
                absent=absent+1
            else:
                present=present+1
        pieattendance1=[absent,present]
        print("Present : ",present,"\nAbsent : ",absent)
        class1date=pd.concat([conroll,conname,condate],axis=1)
        print("Attendance dated on :",indate,"\n\n",class1date)
        plt.pie(pieattendance1, labels=["Absent","Present"],autopct='%1.1f%%',startangle = 90)
        plt.title("Chart of attedance for 1 class on {} date".format(indate))
        plt.show()
    except:print('There is no class on the date for class 1 ',indate)

def dateclass2():
    try:
        dfclass2=pd.read_excel("G:/python project/Class 2.xlsx")
        indate=input("\nDate (DD-MM-YYYY) :")
        conroll=dfclass2['Roll no']
        conname=dfclass2['Name']
        condate=dfclass2[indate]
        absent=0
        present=0
        for i in range(0,30,1):
            if(condate[i]=='A'):
                absent=absent+1
            else:
                present=present+1
        pieattendance2=[absent,present]
        print("Present : ",present,"\nAbsent : ",absent)
        class2date=pd.concat([conroll,conname,condate],axis=1)
        print("Attendance dated on :",indate,"\n\n",class2date)
        plt.pie(pieattendance2, labels=["Absent","Present"],autopct='%1.1f%%',startangle = 90)
        plt.title("Chart of attedance for 2 class on {} date".format(indate))
        plt.show()
    except:print('There is no class on the date for class 2 ',indate)

def dateclass3():
    try:
        dfclass3=pd.read_excel("G:/python project/Class 3.xlsx")
        indate=input("\nDate (DD-MM-YYYY) :")
        conroll=dfclass3['Roll no']
        conname=dfclass3['Name']
        condate=dfclass3[indate]
        absent=0
        present=0
        for i in range(0,30,1):
            if(condate[i]=='A'):
                absent=absent+1
            else:
                present=present+1
        pieattendance3=[absent,present]
        print("Present : ",present,"\nAbsent : ",absent)
        class3date=pd.concat([conroll,conname,condate],axis=1)
        print("Attendance dated on :",indate,"\n\n",class3date)
        plt.pie(pieattendance3, labels=["Absent","Present"],autopct='%1.1f%%',startangle = 90)
        plt.title("Chart of attedance for 3 class on {} date".format(indate))
        plt.show()
    except:print('There is no class on the date for class 3 ',indate)

def dateclass4():
    try:
        dfclass4=pd.read_excel("G:/python project/Class 4.xlsx")
        indate=input("\nDate (DD-MM-YYYY) :")
        conroll=dfclass4['Roll no']
        conname=dfclass4['Name']
        condate=dfclass4[indate]
        absent=0
        present=0
        for i in range(0,30,1):
            if(condate[i]=='A'):
                absent=absent+1
            else:
                present=present+1
        pieattendance4=[absent,present]
        print("Present : ",present,"\nAbsent : ",absent)
        class4date=pd.concat([conroll,conname,condate],axis=1)
        print("Attendance dated on :",indate,"\n\n",class4date)
        plt.pie(pieattendance4, labels=["Absent","Present"],autopct='%1.1f%%',startangle=90)
        plt.title("Chart of attedance for 4 class on {} date".format(indate))
        plt.show()
    except:print('There is no class on the date for class 3 ',indate)

def monday(integerhr,integermin,column):
    if(integerhr>=8 and integerhr <10):
        choice("Class 1",column);
    elif(integerhr>=10 and integerhr<12):
        choice("Free hour",column);
    elif(integerhr>=13 and integerhr<15):
        choice("Class 2",column);
    elif(integerhr>=15 and integerhr<17):
        choice("Class 3",column);
    else:
        choice("Free hour",column);

def tuesday(integerhr,integermin,column):
    column = chr(ord(column) + 1)
    if(integerhr>=8 and integerhr <10):
        choice("Class 3",column);
    elif(integerhr>=10 and integerhr<12):
        choice("Class 2",column);
    elif(integerhr>=13 and integerhr<15):
        choice("Free hour",column);
    elif(integerhr>=15 and integerhr<17):
        choice("Class 1",column);
    else:
        choice("Free hour",column);

def wednesday(integerhr,integermin,column):
    if(integerhr>=8 and integerhr <10):
        choice("Class 4",column);
    elif(integerhr>=10 and integerhr<12):
        choice("Free hour",column);
    elif(integerhr>=13 and integerhr<15):
        choice("Class 1",column);
    elif(integerhr>=15 and integerhr<17):
        choice("Class 3",column);
    else:
        choice("Free hour",column);

def thursday(integerhr,integermin,column):
    if(integerhr>=8 and integerhr <10):
        choice("Class 4",column);
    elif(integerhr>=10 and integerhr<12):
        choice("Class 1",column);
    elif(integerhr>=13 and integerhr<15):
        choice("Class 3",column);
    elif(integerhr>=15 and integerhr<17):
        choice("Free hour",column);
    else:
        choice("Free hour",column);

def friday(integerhr,integermin,column):

    if(integerhr>=8 and integerhr <10):
        choice("Free hour",column);
    elif(integerhr>=10 and integerhr<12):
        choice("Class 2",column);
    elif(integerhr>=13 and integerhr<15):
        choice("Class 4",column);
    elif(integerhr>=15 and integerhr<17):
        choice("Class 3",column);
    else:
        choice("Free hour",column);


def check_day(current_day,integerhr,integermin,column):
    if (current_day == "Monday"):
        s=monday(integerhr,integermin,column);
    elif (current_day == "Tuesday"):
        t=tuesday(integerhr,integermin,column);
    elif (current_day == "Wednesday"):
        u=wednesday(integerhr,integermin,column);
    elif (current_day == "Thursday"):
        v=thursday(integerhr,integermin,column);
    elif (current_day == "Friday"):
        w=friday(integerhr,integermin,column);
    else:
        print("Enjoy yor free time")
    return

def choice(whichclass,column):
    choice=int(input("\n\n1.Attendance report\n2.Attendace for this hour :"+ whichclass +"\n3.Search Attendance by date \n4.Exit\nChoice : "))
    if(choice==1):
        repclass=int(input("\n\n1.Class 1\n2.Class 2\n3.Class 3\n4.Class 4\nClass attendance report : "))
        if repclass==1:
            attendreportclass1();
        elif repclass==2:
            attendreportclass2();
        elif repclass==3:
            attendreportclass3();
        else:
            attendreportclass4();
    elif(choice==2):
        if(whichclass=="Class 1"):
            class1(column);
        elif(whichclass=="Class 2"):
            class2(column);
        elif(whichclass=="Class 3"):
            class3(column);
        elif(whichclass=="Class 4"):
            class4(column);
        elif(whichclass=="Free hour"):
            print("Enjoy your free hour")  
    elif(choice==3):
        dateclass=int(input("\n\n1.Class 1\n2.Class 2\n3.Class 3\n4.Class 4\nClass : "))
        if dateclass==1:
            dateclass1();
        elif dateclass==2:
            dateclass2();
        elif dateclass==3:
            dateclass3();
        else:
            dateclass4();
    else:
        return
            
    
    
    
while(1):
    main=check_day(current_day,integerhr,integermin,column);


