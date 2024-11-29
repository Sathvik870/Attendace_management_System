import tkinter as tk
import datetime
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt

# Create the main window
window = tk.Tk()
window.title("Attendance Management System")

# Set the window size
window.geometry("400x300")

# Function to handle the button click for Class 1 attendance
def class1_attendance():
    column = "S"  # Column for Class 1 attendance
    # Your existing code for class1() goes here
    column="S"
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    # Load the existing workbook
    print("Class 1 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 1.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 1.xlsx")
    print("Attendance registered For class 1")
    return
    
# Function to handle the button click for Class 2 attendance
def class2_attendance():
    column = "O"  # Column for Class 2 attendance
    # Your existing code for class2() goes here
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    # Load the existing workbook
    print("Class 2 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 2.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 2.xlsx")
    print("Attendance registered For class 2")
    return
    
# Function to handle the button click for Class 3 attendance
def class3_attendance():
    column = "W"  # Column for Class 3 attendance
    # Your existing code for class3() goes here
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    # Load the existing workbook
    print("Class 3 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 3.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 3.xlsx")
    print("Attendance registered For class 3")
    return
    
# Function to handle the button click for Class 4 attendance
def class4_attendance():
    column = "O"  # Column for Class 4 attendance
    # Your existing code for class4() goes here
    plusdate=int(current_date_str[8:])
    tday=28
    norm=plusdate-tday
    column = chr(ord(column) + norm)
    print("Class 4 attendance")
    workbook = openpyxl.load_workbook("G:/python project/Class 4.xlsx")

    # Select the active sheet
    sheet = workbook.active

    for i in range(2, 32):
        cell = column + str(i)
        sheet[cell] = input("Roll no "+ str(i-1) +" : ")

    # Save the workbook with modifications
    workbook.save("G:/python project/Class 4.xlsx")
    print("Attendance registered For class 4")
    return
    
# Function to handle the button click for attendance report
def attendance_report():
    # Your existing code for attendreportclass1(), attendreportclass2(), etc. goes here
    def attendreportclass1():
        dfclass1=pd.read_excel("G:/python project/Class 1.xlsx")
        print(dfclass1.to_string())
        return

    def attendreportclass2():
        dfclass2=pd.read_excel("G:/python project/Class 2.xlsx")
        print(dfclass2.to_string())
        return

    def attendreportclass3():
        dfclass3=pd.read_excel("G:/python project/Class 3.xlsx")
        print(dfclass3.to_string())
        return

    def attendreportclass4():
        dfclass4=pd.read_excel("G:/python project/Class 4.xlsx")
        print(dfclass4.to_string())
        return

# Function to handle the button click for searching attendance by date
def search_attendance_by_date():
        # Your existing code for dateclass1(), dateclass2(), etc. goes here
    def dateclass1():
        try:
            dfclass1=pd.read_excel("G:/python project/Class 1.xlsx")
            indate=input("\nDate (DD-MM-YYYY) :")
            conroll=dfclass1['Roll no']
            conname=dfclass1['Name']
            condate=dfclass1[indate]
            absent=0
            present=0
            for i in range(0,30,1):
                if(condate[i]=='A'):
                    absent=absent+1
                else:
                    present=present+1
            pieattendance1=[absent,present]
            print("Present : ",present,"\nAbsent : ",absent)
            class1date=pd.concat([conroll,conname,condate],axis=1)
            print("Attendance dated on :",indate,"\n\n",class1date)
            plt.pie(pieattendance1, labels=["Absent","Present"],autopct='%1.1f%%',startangle = 90)
            plt.title("Chart of attedance for 1 class on {} date".format(indate))
            plt.show()
        except:print('There is no class on the date for class 1 ',indate)

    def dateclass2():
        try:
            dfclass2=pd.read_excel("G:/python project/Class 2.xlsx")
            indate=input("\nDate (DD-MM-YYYY) :")
            conroll=dfclass2['Roll no']
            conname=dfclass2['Name']
            condate=dfclass2[indate]
            absent=0
            present=0
            for i in range(0,30,1):
                if(condate[i]=='A'):
                    absent=absent+1
                else:
                    present=present+1
            pieattendance2=[absent,present]
            print("Present : ",present,"\nAbsent : ",absent)
            class2date=pd.concat([conroll,conname,condate],axis=1)
            print("Attendance dated on :",indate,"\n\n",class2date)
            plt.pie(pieattendance2, labels=["Absent","Present"],autopct='%1.1f%%',startangle = 90)
            plt.title("Chart of attedance for 2 class on {} date".format(indate))
            plt.show()
        except:print('There is no class on the date for class 2 ',indate)

    def dateclass3():
        try:
            dfclass3=pd.read_excel("G:/python project/Class 3.xlsx")
            indate=input("\nDate (DD-MM-YYYY) :")
            conroll=dfclass3['Roll no']
            conname=dfclass3['Name']
            condate=dfclass3[indate]
            absent=0
            present=0
            for i in range(0,30,1):
                if(condate[i]=='A'):
                    absent=absent+1
                else:
                    present=present+1
            pieattendance3=[absent,present]
            print("Present : ",present,"\nAbsent : ",absent)
            class3date=pd.concat([conroll,conname,condate],axis=1)
            print("Attendance dated on :",indate,"\n\n",class3date)
            plt.pie(pieattendance3, labels=["Absent","Present"],autopct='%1.1f%%',startangle = 90)
            plt.title("Chart of attedance for 3 class on {} date".format(indate))
            plt.show()
        except:print('There is no class on the date for class 3 ',indate)

    def dateclass4():
        try:
            dfclass4=pd.read_excel("G:/python project/Class 4.xlsx")
            indate=input("\nDate (DD-MM-YYYY) :")
            conroll=dfclass4['Roll no']
            conname=dfclass4['Name']
            condate=dfclass4[indate]
            absent=0
            present=0
            for i in range(0,30,1):
                if(condate[i]=='A'):
                    absent=absent+1
                else:
                    present=present+1
            pieattendance4=[absent,present]
            print("Present : ",present,"\nAbsent : ",absent)
            class4date=pd.concat([conroll,conname,condate],axis=1)
            print("Attendance dated on :",indate,"\n\n",class4date)
            plt.pie(pieattendance4, labels=["Absent","Present"],autopct='%1.1f%%',startangle=90)
            plt.title("Chart of attedance for 4 class on {} date".format(indate))
            plt.show()
        except:print('There is no class on the date for class 3 ',indate)

# Create buttons
class1_button = tk.Button(window, text="Class 1 Attendance", command=class1_attendance)
class2_button = tk.Button(window, text="Class 2 Attendance", command=class2_attendance)
class3_button = tk.Button(window, text="Class 3 Attendance", command=class3_attendance)
class4_button = tk.Button(window, text="Class 4 Attendance", command=class4_attendance)
report_button = tk.Button(window, text="Attendance Report", command=attendance_report)
search_button = tk.Button(window, text="Search Attendance by Date", command=search_attendance_by_date)

# Place buttons in the window
class1_button.pack()
class2_button.pack()
class3_button.pack()
class4_button.pack()
report_button.pack()
search_button.pack()

# Start the main loop
window.mainloop()
